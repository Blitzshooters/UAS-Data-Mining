"""
Modul untuk menyimpan SEMUA hasil eksperimen (Project 1 & Project 2) ke dalam
SATU file rapi yang siap dipakai untuk laporan:

- Satu file Excel (.xlsx) dengan beberapa sheet terpisah dan terformat:
    * Ringkasan (info dataset, parameter, waktu ekspor)
    * Eksperimen ML Project 1 (tabel hasil + statistik rata2/min/max)
    * Evaluasi LSTM Project 2 (tabel ringkasan hasil)
- Satu file PDF (opsional) yang menggabungkan semua tabel di atas dalam
  format dokumen siap cetak/lampiran laporan.

Dipanggil dari tombol "Export Laporan Lengkap (Excel)" / "... (PDF)" di
aplikasi, bisa dipanggil dari Project 1 maupun Project 2 (mengambil data
dari kedua tab sekaligus jika tersedia).
"""

import datetime

import pandas as pd


def _autofit_and_style(ws, df, header_fill="1F3A5F"):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    header_font = Font(color="FFFFFF", bold=True)
    header_pattern = PatternFill(start_color=header_fill, end_color=header_fill, fill_type="solid")

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_pattern
        cell.alignment = Alignment(horizontal="center")

    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max([len(str(col_name))] + [len(str(v)) for v in df[col_name].astype(str).tolist()])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 3, 10), 45)

    ws.freeze_panes = "A2"


def export_full_report_excel(path, project1_app=None, project2_app=None):
    """
    project1_app: instance Project1Tab (boleh None jika tidak ada data)
    project2_app: instance Project2Tab (boleh None jika tidak ada data)
    """
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        # ---------------------------------------------------- Sheet Ringkasan
        info_rows = [
            ["Laporan Hasil Eksperimen — Sistem Analisis Data dan Deep Learning Terintegrasi"],
            [f"Diekspor pada: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"],
            [""],
        ]
        if project1_app is not None and project1_app.df_raw is not None:
            df1 = project1_app.df_raw
            info_rows += [
                ["PROJECT 1 — Data Preparation & Machine Learning"],
                [f"  Jumlah baris dataset: {df1.shape[0]:,}"],
                [f"  Jumlah kolom dataset: {df1.shape[1]}"],
                [f"  Jumlah eksperimen ML dijalankan: {len(project1_app.experiment_results)}"],
                [""],
            ]
        if project2_app is not None and project2_app.df_raw is not None:
            df2 = project2_app.df_raw
            info_rows += [
                ["PROJECT 2 — Time Series Forecasting (LSTM)"],
                [f"  Jumlah baris dataset: {df2.shape[0]:,}"],
                [f"  Jumlah kolom dataset: {df2.shape[1]}"],
            ]
            if getattr(project2_app, "last_series_meta", None):
                meta = project2_app.last_series_meta
                info_rows += [
                    [f"  Kolom target: {meta.get('target_col')}"],
                    [f"  Filter/grup: {meta.get('filter_col')} = {meta.get('filter_val')}"],
                    [f"  Granularitas series: {meta.get('granularity')}"],
                    [f"  Jumlah titik data series: {meta.get('n_points')}"],
                ]
            if getattr(project2_app, "model", None) is not None:
                m = project2_app.model
                info_rows += [
                    [f"  Sequence Length: {getattr(project2_app, 'seq_len', '-')}"],
                    [f"  Neuron (Hidden Units): {m.n_hidden}"],
                    [f"  Dropout: {m.dropout}"],
                    [f"  Jumlah epoch dilatih: {len(m.history.get('loss', []))}"],
                ]
        info_df = pd.DataFrame(info_rows)
        info_df.to_excel(writer, sheet_name="Ringkasan", header=False, index=False)

        # ---------------------------------------------------- Sheet Project 1
        if project1_app is not None and project1_app.experiment_results:
            results_df = pd.DataFrame(project1_app.experiment_results)
            results_df.to_excel(writer, sheet_name="P1_Eksperimen_ML", index=False)
            _autofit_and_style(writer.sheets["P1_Eksperimen_ML"], results_df)

            # statistik rata2/min/max
            stats_rows = []
            for model in ["RF", "LR"]:
                cols = {"Precision": f"Prec_{model}", "Recall": f"Rec_{model}",
                        "F1 Score": f"F1_{model}", "Accuracy": f"Acc_{model}"}
                for stat_name, fn in [("Rata-rata", "mean"), ("Minimum", "min"), ("Maksimum", "max")]:
                    row = {"Model": model, "Statistik": stat_name}
                    for k, v in cols.items():
                        row[k] = round(getattr(results_df[v], fn)(), 3)
                    stats_rows.append(row)
            stats_df = pd.DataFrame(stats_rows)
            stats_df.to_excel(writer, sheet_name="P1_Statistik_Performa", index=False)
            _autofit_and_style(writer.sheets["P1_Statistik_Performa"], stats_df)

        # missing value & outlier tables (kalau dataset ada)
        if project1_app is not None and project1_app.df_raw is not None:
            df1 = project1_app.df_raw
            import numpy as np
            miss_count = df1.isna().sum()
            miss_pct = (miss_count / len(df1) * 100).round(2)
            miss_df = pd.DataFrame({"Kolom": df1.columns, "Jumlah Missing": miss_count.values,
                                     "Persentase (%)": miss_pct.values})
            miss_df = miss_df[miss_df["Jumlah Missing"] > 0].sort_values("Persentase (%)", ascending=False)
            if len(miss_df) > 0:
                miss_df.to_excel(writer, sheet_name="P1_Missing_Value", index=False)
                _autofit_and_style(writer.sheets["P1_Missing_Value"], miss_df)

            num_cols = df1.select_dtypes(include=np.number).columns
            outlier_rows = []
            for col in num_cols:
                s = df1[col].dropna()
                if len(s) == 0:
                    continue
                q1, q3 = s.quantile(0.25), s.quantile(0.75)
                iqr = q3 - q1
                lower, upper = q1 - 1.5 * iqr, q3 + 1.5 * iqr
                n_out = ((s < lower) | (s > upper)).sum()
                outlier_rows.append({"Kolom": col, "Jumlah Outlier (IQR)": int(n_out),
                                      "Lower Bound": round(lower, 2), "Upper Bound": round(upper, 2)})
            if outlier_rows:
                outlier_df = pd.DataFrame(outlier_rows).sort_values("Jumlah Outlier (IQR)", ascending=False)
                outlier_df.to_excel(writer, sheet_name="P1_Outlier_Analysis", index=False)
                _autofit_and_style(writer.sheets["P1_Outlier_Analysis"], outlier_df)

        # ---------------------------------------------------- Sheet Project 2
        if project2_app is not None and getattr(project2_app, "eval_results", None):
            eval_df = pd.DataFrame(project2_app.eval_results)
            eval_df.to_excel(writer, sheet_name="P2_Evaluasi_LSTM", index=False)
            _autofit_and_style(writer.sheets["P2_Evaluasi_LSTM"], eval_df)

        if project2_app is not None and getattr(project2_app, "model", None) is not None:
            hist = project2_app.model.history
            hist_df = pd.DataFrame({
                "Epoch": list(range(1, len(hist.get("loss", [])) + 1)),
                "Train Loss": hist.get("loss", []),
                "Validation Loss": hist.get("val_loss", []),
            })
            if len(hist_df) > 0:
                hist_df.to_excel(writer, sheet_name="P2_Riwayat_Loss", index=False)
                _autofit_and_style(writer.sheets["P2_Riwayat_Loss"], hist_df)

        if project2_app is not None and hasattr(project2_app, "last_pred"):
            pred_df = pd.DataFrame({
                "Aktual": project2_app.last_actual,
                "Prediksi": project2_app.last_pred,
                "Error": project2_app.last_pred - project2_app.last_actual,
            })
            pred_df.to_excel(writer, sheet_name="P2_Prediksi_vs_Aktual", index=False)
            _autofit_and_style(writer.sheets["P2_Prediksi_vs_Aktual"], pred_df)


def export_full_report_pdf(path, project1_app=None, project2_app=None):
    from matplotlib.backends.backend_pdf import PdfPages
    import matplotlib.pyplot as plt

    def render_table_page(pdf, title, df, note=None):
        n_rows_per_page = 22
        n_pages = max(1, (len(df) + n_rows_per_page - 1) // n_rows_per_page)
        for p in range(n_pages):
            chunk = df.iloc[p * n_rows_per_page:(p + 1) * n_rows_per_page]
            fig, ax = plt.subplots(figsize=(11.7, 8.3))
            ax.axis("off")
            page_title = title if n_pages == 1 else f"{title} (hal. {p+1}/{n_pages})"
            ax.set_title(page_title, fontsize=13, fontweight="bold", pad=20)
            if note and p == 0:
                ax.text(0, 1.02, note, transform=ax.transAxes, fontsize=8, color="gray")
            tbl = ax.table(cellText=chunk.values, colLabels=chunk.columns, loc="center", cellLoc="center")
            tbl.auto_set_font_size(False)
            tbl.set_fontsize(7)
            tbl.scale(1, 1.3)
            pdf.savefig(fig)
            plt.close(fig)

    with PdfPages(path) as pdf:
        # cover
        fig, ax = plt.subplots(figsize=(11.7, 8.3))
        ax.axis("off")
        ax.text(0.5, 0.7, "LAPORAN HASIL EKSPERIMEN", ha="center", fontsize=20, fontweight="bold")
        ax.text(0.5, 0.62, "Sistem Analisis Data dan Deep Learning Terintegrasi", ha="center", fontsize=14)
        ax.text(0.5, 0.5, f"Diekspor pada: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                ha="center", fontsize=10, color="gray")
        pdf.savefig(fig)
        plt.close(fig)

        if project1_app is not None and project1_app.experiment_results:
            results_df = pd.DataFrame(project1_app.experiment_results)
            render_table_page(pdf, "PROJECT 1 — Tabel Hasil Eksperimen Machine Learning", results_df)

            stats_rows = []
            for model in ["RF", "LR"]:
                cols = {"Precision": f"Prec_{model}", "Recall": f"Rec_{model}",
                        "F1 Score": f"F1_{model}", "Accuracy": f"Acc_{model}"}
                for stat_name, fn in [("Rata-rata", "mean"), ("Minimum", "min"), ("Maksimum", "max")]:
                    row = {"Model": model, "Statistik": stat_name}
                    for k, v in cols.items():
                        row[k] = round(getattr(results_df[v], fn)(), 3)
                    stats_rows.append(row)
            render_table_page(pdf, "PROJECT 1 — Statistik Performa Model", pd.DataFrame(stats_rows))

        if project2_app is not None and getattr(project2_app, "eval_results", None):
            eval_df = pd.DataFrame(project2_app.eval_results)
            note = None
            if getattr(project2_app, "last_series_meta", None):
                meta = project2_app.last_series_meta
                note = (f"Target: {meta.get('target_col')} | Filter: {meta.get('filter_col')}="
                        f"{meta.get('filter_val')} | Granularitas: {meta.get('granularity')}")
            render_table_page(pdf, "PROJECT 2 — Tabel Ringkasan Evaluasi LSTM", eval_df, note=note)

        if project2_app is not None and getattr(project2_app, "model", None) is not None:
            hist = project2_app.model.history
            if hist.get("loss"):
                fig, ax = plt.subplots(figsize=(11.7, 6))
                ax.plot(hist["loss"], label="Training Loss", color="#337ab7")
                ax.plot(hist["val_loss"], label="Validation Loss", color="#d9534f")
                ax.set_xlabel("Epoch"); ax.set_ylabel("MSE Loss")
                ax.set_title("PROJECT 2 — Loss Training vs Validation")
                ax.legend()
                pdf.savefig(fig)
                plt.close(fig)

        if project2_app is not None and hasattr(project2_app, "last_pred"):
            fig, ax = plt.subplots(figsize=(11.7, 6))
            ax.plot(project2_app.last_actual, label="Aktual", color="#337ab7")
            ax.plot(project2_app.last_pred, label="Prediksi", color="#d9534f", linestyle="--")
            ax.set_title("PROJECT 2 — Prediksi vs Aktual (Test Set)")
            ax.legend()
            pdf.savefig(fig)
            plt.close(fig)


def run_export_dialog(project1_app=None, project2_app=None, fmt="xlsx"):
    """Dipanggil dari tombol UI. Membuka dialog Save As lalu mengekspor."""
    from tkinter import filedialog, messagebox

    has_p1 = project1_app is not None and project1_app.experiment_results
    has_p2 = project2_app is not None and getattr(project2_app, "eval_results", None)
    if not has_p1 and not has_p2:
        messagebox.showwarning(
            "Peringatan",
            "Belum ada hasil eksperimen dari Project 1 maupun Project 2.\n"
            "Jalankan minimal satu eksperimen ML (Project 1) atau training+evaluasi LSTM "
            "(Project 2) sebelum mengekspor laporan lengkap."
        )
        return

    if fmt == "xlsx":
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")],
                                             initialfile="Laporan_Hasil_Eksperimen_Lengkap.xlsx")
        if not path:
            return
        try:
            export_full_report_excel(path, project1_app, project2_app)
            messagebox.showinfo("Export Berhasil", f"Laporan lengkap berhasil disimpan ke:\n{path}\n\n"
                                 "File ini berisi semua sheet (ringkasan, tabel eksperimen, statistik, "
                                 "evaluasi LSTM) dalam satu file siap dilampirkan ke laporan.")
        except Exception as e:
            messagebox.showerror("Error", f"Gagal mengekspor laporan lengkap:\n{e}")
    else:
        path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF", "*.pdf")],
                                             initialfile="Laporan_Hasil_Eksperimen_Lengkap.pdf")
        if not path:
            return
        try:
            export_full_report_pdf(path, project1_app, project2_app)
            messagebox.showinfo("Export Berhasil", f"Laporan lengkap berhasil disimpan ke:\n{path}")
        except Exception as e:
            messagebox.showerror("Error", f"Gagal mengekspor laporan PDF:\n{e}")
